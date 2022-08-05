from flask import Flask
import tempfile

from flask_restful import Api, Resource 

from flask import request, jsonify
from tablib import Dataset
import os
import time
import urllib.parse
from werkzeug.utils import secure_filename
import pandas
from openpyxl import load_workbook


app = Flask (__name__)

api= Api(app)



@app.route('/upload', methods=['POST'])
def upload_file():
    category = request.args.get('category')
    # I used form data type which means there is a
    # "Content-Type: application/x-www-form-urlencoded"
    # header in my request
    #raw_data = request.files['myfile'].read()  # In form data, I used "myfile" as key.
    #dataset = Dataset().load(raw_data)
    #folder = os.path.join(app.instance_path, 'uploads')
    # os.makedirs(folder)
    file= request.files['myfile']

   #file.save(os.path.join(folder, "Liste_colonnes_obligatoires.xlsx"))

    fl_secure = secure_filename(file.filename)

    #file.save('C:\\Users\\smala\\Desktop\\res\\' + fl_secure)

    wb = load_workbook(file)
    ws = wb.active

    
    MyExcelFileColumns=[]
            
    for i in range(1,ws.max_column+1):
        cell_obj = ws.cell(row=1, column=i)
        MyExcelFileColumns.append(cell_obj.value)
    
   
    bo=False
    #data=jsonify(dataset.export('json'))
    #print(str(data)+"salem")
    MyNewList=["Invoice","MasterData","ProfitLoss","EnironmentalPerformance","HumanResource"]
    if category in MyNewList:
        bo=True

    
    

    thisdict = {"InvoiceList":["Demande d'énergie primaire","Total Electricité","Taxes et contributions","Montant TTC à payer","Consommation","Conso Chauffage","Conso CVC","Conso Froid","Conso Eclairage","Conso auxiliaire","Emission GES: Scope 1 + 2","Energies renouvelables"],
    "MasterDataList":["Property","Region","Address","Floor Area (m2)","Construction Year","Type Du Batiment"],
    "ProfitLossList":["Revenue","Total Operating Expenses","Accounts Receivable","Inventory","Property & Equipment","Total Assets","CapEx"],
    "EnironmentalPerformanceList":['Area','Type Emission','Type Conso','Total','Unit'],
    "HumanResourceList":["Climate change mitigation","Climate change adaptation","Sustainable use and protection of water and marine resources","Transition to circular economy","Pollution prevention and control","Protection and restoration of biodiversity and ecosystems"]
    }


    print(MyExcelFileColumns)
    toScrapList=[]

    DoesNotContainList=[]
    ContainList=[]
    temp=category+"List"
    bouli=True
    for ahla in list(thisdict):
        if(temp==ahla):
            
            toScrapList=thisdict[ahla]


    if bo:
        for k in toScrapList:
            for l in MyExcelFileColumns:
                if k==l:
                    ContainList.append(k)
    
    print(ContainList)   

    DoesNotContainList=list(set(toScrapList) - set(ContainList))
    print(DoesNotContainList)

    thisdict2={
        "Does not contain list":DoesNotContainList
    }
                 




    
    return (thisdict2)
    

    





if __name__ == "__main__" : 
    app.run(debug=True)

